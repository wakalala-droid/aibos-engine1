# ════════════════════════════════════════════════════════
# AI-BOS — Financial Intelligence Engine  (Week 3 update)
# New: export_excel_report · chat persistence · email send
# ════════════════════════════════════════════════════════

import io
import os
import json
import smtplib
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text      import MIMEText
from email.mime.base      import MIMEBase
from email                import encoders

import numpy  as np
import pandas as pd
from openpyxl                        import Workbook
from openpyxl.styles                 import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.styles.numbers         import FORMAT_NUMBER_COMMA_SEPARATED1, FORMAT_PERCENTAGE_00
from openpyxl.utils                  import get_column_letter
from groq import Groq

# ── AI CLIENT ─────────────────────────────────────────────
client_ai = Groq(api_key=os.environ.get("GROQ_API_KEY") or "placeholder")

# ═════════════════════════════════════════════════════════
# ORIGINAL ENGINE FUNCTIONS (unchanged)
# ═════════════════════════════════════════════════════════

def analyse_pnl(df):
    total_rev    = df["revenue"].sum()
    total_cost   = df["costs"].sum()
    total_profit = total_rev - total_cost
    avg_margin   = round(df["margin_pct"].mean(), 1)
    best_idx     = df["profit"].idxmax()
    return {
        "total_revenue": total_rev,
        "total_costs":   total_cost,
        "total_profit":  total_profit,
        "avg_margin":    avg_margin,
        "best_month":    df.loc[best_idx, "month"],
        "worst_month":   df.loc[df["profit"].idxmin(), "month"],
    }


def forecast_cashflow(df, current_cash=50000, months_ahead=3):
    recent_profit = df.tail(3)["profit"].mean()
    running_cash  = current_cash
    forecast      = []
    for m in range(1, months_ahead + 1):
        running_cash += recent_profit
        forecast.append({
            "month_ahead":    m,
            "projected_cash": round(running_cash),
            "status": "⚠ NEGATIVE" if running_cash < 0 else "✓ Positive",
        })
    return forecast


def detect_variances(df, threshold=0.10):
    alerts = []
    for i in range(1, len(df)):
        prev_rev = df.iloc[i-1]["revenue"]
        curr_rev = df.iloc[i]["revenue"]
        pct_chg  = ((curr_rev - prev_rev) / prev_rev) * 100
        if abs(pct_chg) > threshold * 100:
            alerts.append({
                "month":      df.iloc[i]["month"],
                "change_pct": round(pct_chg, 1),
                "direction":  "drop" if pct_chg < 0 else "spike",
            })
    return alerts


def health_score(pnl, alerts):
    margin_pts  = min(pnl["avg_margin"] * 1.5, 50)
    trend_pts   = 30 if pnl["total_revenue"] - pnl["total_costs"] > 0 else 0
    alert_pen   = len(alerts) * 8
    score       = max(0, min(100, margin_pts + trend_pts - alert_pen))
    label       = ("Critical" if score < 30 else "At Risk" if score < 55
                   else "Healthy" if score < 80 else "Excellent")
    return round(score), label


def get_structured_analysis(pnl, alerts):
    prompt = f"""You are an expert CFO advising a small African business.
Revenue: K{pnl['total_revenue']:,} | Costs: K{pnl['total_costs']:,} | Profit: K{pnl['total_profit']:,}
Margin: {pnl['avg_margin']}% | Best: {pnl['best_month']} | Worst: {pnl['worst_month']} | Alerts: {len(alerts)}

Return ONLY valid JSON, no other text:
[{{"title":"...","recommendation":"...","priority":"high/medium/low"}},{{"title":"...","recommendation":"...","priority":"high/medium/low"}},{{"title":"...","recommendation":"...","priority":"high/medium/low"}}]"""
    raw = client_ai.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        timeout=30,
    ).choices[0].message.content
    try:
        return json.loads(raw)
    except Exception:
        return [{"title": "Analysis", "recommendation": raw, "priority": "medium"}]


def forecast_revenue(df: pd.DataFrame, months_ahead: int = 6) -> dict:
    revenue = pd.to_numeric(df["revenue"], errors="coerce").fillna(0).values
    n = len(revenue)
    if n < 3:
        return {"error": "Need at least 3 months of data"}

    x = np.arange(n, dtype=float)
    slope, intercept = np.polyfit(x, revenue, 1)
    y_hat     = slope * x + intercept
    residuals = revenue - y_hat
    std_err   = float(np.std(residuals))
    ss_res    = float(np.sum(residuals ** 2))
    ss_tot    = float(np.sum((revenue - np.mean(revenue)) ** 2))
    r2        = max(0.0, 1 - ss_res / ss_tot) if ss_tot > 0 else 0.0
    confidence   = int(min(100, round(r2 * 100)))
    last_rev     = float(revenue[-1]) if revenue[-1] > 0 else 1.0
    growth_rate  = round((slope / last_rev) * 100, 2)
    trend        = ("upward" if slope > last_rev * 0.01
                    else "downward" if slope < -last_rev * 0.01 else "flat")

    forecast_pts = []
    for i in range(1, months_ahead + 1):
        predicted = max(0.0, slope * (n - 1 + i) + intercept)
        forecast_pts.append({
            "month":     f"Month +{i}",
            "predicted": round(predicted),
            "low":       round(max(0.0, predicted - std_err)),
            "high":      round(predicted + std_err),
        })

    try:
        next_m = forecast_pts[0]["predicted"]
        last_m = forecast_pts[-1]["predicted"]
        prompt = (f"CFO advisor. 2 sentences. Trend:{trend} Growth:{growth_rate}%/mo "
                  f"Confidence:{confidence}/100 Next:K{next_m:,} End:K{last_m:,}. Plain English.")
        ai_exp = client_ai.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            timeout=20,
        ).choices[0].message.content.strip()
    except Exception:
        ai_exp = f"Revenue projected to {trend} at {abs(growth_rate)}%/mo. Next month: K{forecast_pts[0]['predicted']:,}."

    return {
        "forecast": forecast_pts, "trend": trend, "growth_rate": growth_rate,
        "confidence": confidence, "r_squared": round(r2, 3),
        "ai_explanation": ai_exp, "std_err": round(std_err),
    }


def detect_anomalies(df: pd.DataFrame, z_threshold: float = 2.0) -> list:
    if len(df) < 4:
        return []
    anomalies = []
    metrics   = {"revenue": "Revenue", "costs": "Costs", "margin_pct": "Profit Margin"}
    order_map = {"critical": 0, "high": 1, "medium": 2, "low": 3}

    for col, label in metrics.items():
        if col not in df.columns:
            continue
        series = pd.to_numeric(df[col], errors="coerce").fillna(0)
        mean, std = series.mean(), series.std()
        if std == 0:
            continue
        for i in range(1, len(df)):
            z = (series.iloc[i] - mean) / std
            if abs(z) < z_threshold:
                continue
            prev       = series.iloc[i - 1]
            pct_change = ((series.iloc[i] - prev) / prev * 100) if prev != 0 else 0.0
            direction  = "drop" if z < 0 else "spike"
            severity   = ("critical" if abs(z) >= 3.0 else "high" if abs(z) >= 2.5
                          else "medium" if abs(z) >= 2.0 else "low")
            try:
                rc_prompt = (f"CFO analyst. 1-2 sentences. Root cause of: {label} {direction} "
                             f"{abs(round(pct_change,1))}% in {df.iloc[i]['month']} "
                             f"({round(abs(z),2)}σ, {severity}). No preamble.")
                root_cause = client_ai.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[{"role": "user", "content": rc_prompt}],
                    timeout=20,
                ).choices[0].message.content.strip()
            except Exception:
                root_cause = f"{label} {direction} of {abs(round(pct_change,1))}% — manual review recommended."

            anomalies.append({
                "month": str(df.iloc[i]["month"]), "metric": label,
                "type": f"{col}_{direction}", "direction": direction,
                "change_pct": round(pct_change, 1), "z_score": round(abs(z), 2),
                "severity": severity, "root_cause": root_cause,
            })

    anomalies.sort(key=lambda a: order_map.get(a["severity"], 4))
    return anomalies


def calculate_breakeven(df: pd.DataFrame, fixed_cost_pct: float = 0.40) -> dict:
    avg_revenue    = float(df["revenue"].mean())
    avg_costs      = float(df["costs"].mean())
    avg_profit     = float(df["profit"].mean()) if "profit" in df.columns else avg_revenue - avg_costs
    fixed_costs    = avg_costs * fixed_cost_pct
    variable_costs = avg_costs * (1 - fixed_cost_pct)
    vcr            = variable_costs / avg_revenue if avg_revenue > 0 else 0
    cmr            = 1 - vcr
    breakeven_revenue = fixed_costs / cmr if cmr > 0 else avg_costs
    margin_of_safety  = avg_revenue - breakeven_revenue
    mos_pct           = (margin_of_safety / avg_revenue * 100) if avg_revenue > 0 else 0
    breakeven_pct     = (breakeven_revenue / avg_revenue * 100) if avg_revenue > 0 else 100
    months_to_profit  = None
    if avg_profit < 0 and avg_revenue > 0:
        months_to_profit = round((breakeven_revenue - avg_revenue) / (avg_revenue * 0.05), 1)
    scenarios = []
    for pct in [5, 10, 15, 20]:
        nf = fixed_costs * (1 + pct / 100)
        nb = nf / cmr if cmr > 0 else avg_costs
        scenarios.append({
            "cost_increase_pct": pct, "new_breakeven": round(nb),
            "margin_of_safety": round(avg_revenue - nb),
            "status": "safe" if avg_revenue - nb > 0 else "at risk",
        })
    try:
        status = "above" if avg_revenue >= breakeven_revenue else "below"
        be_prompt = (f"CFO advisor. 2 sentences. Revenue:K{avg_revenue:,.0f} "
                     f"Breakeven:K{breakeven_revenue:,.0f} Status:{status} "
                     f"Safety:{mos_pct:.1f}% Fixed:{fixed_cost_pct*100:.0f}% of costs.")
        ai_insight = client_ai.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": be_prompt}],
            timeout=20,
        ).choices[0].message.content.strip()
    except Exception:
        status = "above" if avg_revenue >= breakeven_revenue else "below"
        ai_insight = (f"Business is {status} breakeven of K{breakeven_revenue:,.0f}/month. "
                      f"Safety margin: {mos_pct:.1f}%.")

    return {
        "breakeven_revenue": round(breakeven_revenue), "current_avg_revenue": round(avg_revenue),
        "breakeven_pct": round(breakeven_pct, 1), "contribution_margin_ratio": round(cmr * 100, 1),
        "fixed_costs": round(fixed_costs), "variable_costs": round(variable_costs),
        "margin_of_safety": round(margin_of_safety), "margin_of_safety_pct": round(mos_pct, 1),
        "months_to_profit": months_to_profit, "scenarios": scenarios, "ai_insight": ai_insight,
    }


# ═════════════════════════════════════════════════════════
# WEEK 3 — FEATURE 1: FORMATTED EXCEL EXPORT
# ═════════════════════════════════════════════════════════

# ── Style helpers ──────────────────────────────────────────
_NAVY   = "1E3A5F"
_BLUE   = "2563EB"
_TEAL   = "0891B2"
_GREEN  = "059669"
_RED    = "DC2626"
_AMBER  = "D97706"
_LGREY  = "F1F5F9"
_WHITE  = "FFFFFF"
_DARK   = "0F172A"


def _hdr(ws, row, col, value, bg=_NAVY, fg=_WHITE, bold=True, size=10, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return c


def _cell(ws, row, col, value, bold=False, color=_DARK, fmt=None, bg=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=color, size=9)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    return c


def _thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def _sheet_summary(wb, pnl, score, label, alerts, runway_months):
    ws = wb.create_sheet("Executive Summary", 0)
    ws.sheet_view.showGridLines = False

    # Title band
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "AI-BOS · Executive Intelligence Report"
    t.font      = Font(name="Arial", bold=True, size=14, color=_WHITE)
    t.fill      = PatternFill("solid", start_color=_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value     = f"Generated {datetime.date.today().strftime('%d %B %Y')}"
    sub.font      = Font(name="Arial", size=9, color="64748B")
    sub.fill      = PatternFill("solid", start_color="EFF6FF")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Health Score banner
    ws.merge_cells("A4:C5")
    score_color = (_GREEN if label == "Excellent" else _BLUE if label == "Healthy"
                   else _AMBER if label == "At Risk" else _RED)
    h = ws["A4"]
    h.value     = f"Health Score: {score}/100 — {label}"
    h.font      = Font(name="Arial", bold=True, size=16, color=score_color)
    h.alignment = Alignment(horizontal="center", vertical="center")

    # KPIs
    kpi_data = [
        ("Total Revenue",  pnl["total_revenue"],  '#,##0.00', _BLUE),
        ("Total Costs",    pnl["total_costs"],    '#,##0.00', _RED),
        ("Net Profit",     pnl["total_profit"],   '#,##0.00',
         _GREEN if pnl["total_profit"] >= 0 else _RED),
        ("Avg Margin %",   pnl["avg_margin"],     '0.0"%"',    _TEAL),
        ("Cash Runway",    f"{runway_months:.1f} mo", None,  "64748B"),
        ("Variance Alerts", len(alerts),           None,
         _AMBER if alerts else _GREEN),
        ("Best Month",     pnl["best_month"],     None,       _GREEN),
        ("Worst Month",    pnl["worst_month"],    None,       _RED),
    ]
    _hdr(ws, 7, 1, "METRIC",     bg=_NAVY, size=9)
    _hdr(ws, 7, 2, "VALUE",      bg=_NAVY, size=9)
    _hdr(ws, 7, 3, "STATUS",     bg=_NAVY, size=9)

    for r, (name, val, fmt, col) in enumerate(kpi_data, start=8):
        bg = _LGREY if r % 2 == 0 else _WHITE
        _cell(ws, r, 1, name, bold=True, bg=bg)
        c = _cell(ws, r, 2, val, bold=True, color=col, bg=bg, align="right")
        if fmt:
            c.number_format = fmt
        bar = "█" * min(20, int(abs(val) / max(pnl["total_revenue"] / 20, 1))) if isinstance(val, (int, float)) else ""
        _cell(ws, r, 3, bar, color=col, bg=bg)
        for col_idx in range(1, 4):
            ws.cell(r, col_idx).border = _thin_border()

    # Alerts table
    if alerts:
        alert_start = 8 + len(kpi_data) + 2
        _hdr(ws, alert_start, 1, "VARIANCE ALERTS", bg=_RED, size=9)
        _hdr(ws, alert_start, 2, "Month",           bg=_RED, size=9)
        _hdr(ws, alert_start, 3, "Change %",        bg=_RED, size=9)
        for r, a in enumerate(alerts, start=alert_start + 1):
            col = _RED if a["direction"] in ("drop", "down") else _AMBER
            bg  = _LGREY if r % 2 == 0 else _WHITE
            _cell(ws, r, 1, a.get("type", "variance").replace("_", " ").title(), bg=bg)
            _cell(ws, r, 2, str(a["month"]), bg=bg)
            _cell(ws, r, 3, a["change_pct"], bold=True, color=col, fmt="0.0", bg=bg, align="right")

    _set_col_widths(ws, {"A": 22, "B": 18, "C": 26, "D": 12, "E": 12, "F": 12})
    return ws


def _sheet_pnl(wb, df, pnl):
    ws = wb.create_sheet("P&L Data")
    ws.sheet_view.showGridLines = False

    headers = ["Month", "Revenue (K)", "Costs (K)", "Profit (K)", "Margin %"]
    for c, h in enumerate(headers, start=1):
        _hdr(ws, 1, c, h, bg=_NAVY, size=10)
        ws.row_dimensions[1].height = 20

    cols = ["month", "revenue", "costs", "profit", "margin_pct"]
    fmts = [None, '#,##0', '#,##0', '#,##0', '0.0"%"']
    for r, (_, row) in enumerate(df.iterrows(), start=2):
        bg = _LGREY if r % 2 == 0 else _WHITE
        for c, (col, fmt) in enumerate(zip(cols, fmts), start=1):
            val = row[col] if col in row else ""
            profit_col = (_GREEN if col == "profit" and float(val or 0) >= 0
                          else _RED if col == "profit" else _DARK)
            cell = _cell(ws, r, c, val, color=profit_col, fmt=fmt, bg=bg,
                         align="right" if c > 1 else "left")
            cell.border = _thin_border()

    # Totals row
    tr = len(df) + 2
    _cell(ws, tr, 1, "TOTALS", bold=True, bg=_NAVY, color=_WHITE)
    total_cells = {2: pnl["total_revenue"], 3: pnl["total_costs"], 4: pnl["total_profit"]}
    for c, val in total_cells.items():
        col = _GREEN if c == 4 and val >= 0 else _RED if c == 4 else _WHITE
        cell = _cell(ws, tr, c, val, bold=True, color=col, fmt='#,##0', bg=_NAVY, align="right")
        cell.border = _thin_border()
    _cell(ws, tr, 5, f"{pnl['avg_margin']}%", bold=True, color=_WHITE,
          fmt='0.0"%"', bg=_NAVY, align="right")

    _set_col_widths(ws, {"A": 14, "B": 16, "C": 16, "D": 16, "E": 12})
    _freeze(ws)
    return ws


def _sheet_forecast(wb, fc_data):
    if not fc_data or "error" in fc_data:
        return
    ws = wb.create_sheet("Revenue Forecast")
    ws.sheet_view.showGridLines = False

    # Trend banner
    t_color = (_GREEN if fc_data["trend"] == "upward" else
               _RED if fc_data["trend"] == "downward" else _AMBER)
    ws.merge_cells("A1:E1")
    banner = ws["A1"]
    banner.value     = (f"Trend: {fc_data['trend'].upper()} | "
                        f"Growth: {fc_data['growth_rate']:+.1f}%/mo | "
                        f"Confidence: {fc_data['confidence']}/100")
    banner.font      = Font(name="Arial", bold=True, size=11, color=_WHITE)
    banner.fill      = PatternFill("solid", start_color=t_color)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(["Period", "Forecast (K)", "Low Case (K)", "High Case (K)", "Range (K)"], start=1):
        _hdr(ws, 3, c, h, bg=_NAVY, size=9)

    for r, pt in enumerate(fc_data["forecast"], start=4):
        bg = _LGREY if r % 2 == 0 else _WHITE
        _cell(ws, r, 1, pt["month"], bg=bg)
        _cell(ws, r, 2, pt["predicted"], bold=True, color=_BLUE, fmt='#,##0', bg=bg, align="right")
        _cell(ws, r, 3, pt["low"],       color="64748B", fmt='#,##0', bg=bg, align="right")
        _cell(ws, r, 4, pt["high"],      color="64748B", fmt='#,##0', bg=bg, align="right")
        _cell(ws, r, 5, pt["high"] - pt["low"], fmt='#,##0', bg=bg, align="right")
        for c in range(1, 6):
            ws.cell(r, c).border = _thin_border()

    # AI narrative
    note_row = len(fc_data["forecast"]) + 6
    ws.merge_cells(f"A{note_row}:E{note_row}")
    n = ws[f"A{note_row}"]
    n.value     = f"AI Narrative: {fc_data.get('ai_explanation', '')}"
    n.font      = Font(name="Arial", italic=True, size=9, color="475569")
    n.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height = 42

    _set_col_widths(ws, {"A": 14, "B": 16, "C": 16, "D": 16, "E": 14})
    _freeze(ws)


def _sheet_anomalies(wb, anomalies):
    if not anomalies:
        return
    ws = wb.create_sheet("Anomaly Intelligence")
    ws.sheet_view.showGridLines = False

    for c, h in enumerate(["Severity", "Month", "Metric", "Direction", "Change %", "Z-Score", "Root Cause"], start=1):
        _hdr(ws, 1, c, h, bg=_NAVY, size=9)

    sev_colors = {"critical": _RED, "high": _AMBER, "medium": _BLUE, "low": _GREEN}
    for r, a in enumerate(anomalies, start=2):
        bg    = _LGREY if r % 2 == 0 else _WHITE
        color = sev_colors.get(a["severity"], _BLUE)
        _cell(ws, r, 1, a["severity"].upper(), bold=True, color=color, bg=bg)
        _cell(ws, r, 2, str(a["month"]),  bg=bg)
        _cell(ws, r, 3, a["metric"],      bg=bg)
        _cell(ws, r, 4, a["direction"],   bg=bg)
        _cell(ws, r, 5, a["change_pct"], bold=True, color=color, fmt="0.0", bg=bg, align="right")
        _cell(ws, r, 6, a["z_score"],    fmt="0.00", bg=bg, align="right")
        rc = _cell(ws, r, 7, a["root_cause"], bg=bg)
        rc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 32
        for c in range(1, 8):
            ws.cell(r, c).border = _thin_border()

    _set_col_widths(ws, {"A": 12, "B": 12, "C": 16, "D": 10, "E": 10, "F": 10, "G": 50})
    _freeze(ws)


def _sheet_breakeven(wb, be):
    if not be:
        return
    ws = wb.create_sheet("Breakeven Analysis")
    ws.sheet_view.showGridLines = False

    above  = be["current_avg_revenue"] >= be["breakeven_revenue"]
    status = "ABOVE BREAKEVEN ✓" if above else "BELOW BREAKEVEN ✗"
    color  = _GREEN if above else _RED

    ws.merge_cells("A1:D1")
    b = ws["A1"]
    b.value     = status
    b.font      = Font(name="Arial", bold=True, size=14, color=_WHITE)
    b.fill      = PatternFill("solid", start_color=color)
    b.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    rows = [
        ("Breakeven Revenue (K/mo)",  be["breakeven_revenue"],    '#,##0', color),
        ("Current Avg Revenue (K/mo)",be["current_avg_revenue"],  '#,##0', _BLUE),
        ("Fixed Costs (K/mo)",        be["fixed_costs"],          '#,##0', _AMBER),
        ("Variable Costs (K/mo)",     be["variable_costs"],       '#,##0', _AMBER),
        ("Contribution Margin Ratio", be["contribution_margin_ratio"], '0.0"%"', _TEAL),
        ("Margin of Safety (K)",      be["margin_of_safety"],     '#,##0',
         _GREEN if be["margin_of_safety"] >= 0 else _RED),
        ("Margin of Safety %",        be["margin_of_safety_pct"], '0.0"%"',
         _GREEN if be["margin_of_safety_pct"] >= 0 else _RED),
    ]
    _hdr(ws, 3, 1, "METRIC",  bg=_NAVY, size=9)
    _hdr(ws, 3, 2, "VALUE",   bg=_NAVY, size=9)

    for r, (name, val, fmt, col) in enumerate(rows, start=4):
        bg = _LGREY if r % 2 == 0 else _WHITE
        _cell(ws, r, 1, name, bold=True, bg=bg)
        _cell(ws, r, 2, val, bold=True, color=col, fmt=fmt, bg=bg, align="right")
        for c in (1, 2):
            ws.cell(r, c).border = _thin_border()

    # What-if table
    _hdr(ws, 13, 1, "WHAT-IF SCENARIOS",   bg=_RED,  size=9)
    _hdr(ws, 13, 2, "Cost Increase",        bg=_RED,  size=9)
    _hdr(ws, 13, 3, "New Breakeven (K)",    bg=_RED,  size=9)
    _hdr(ws, 13, 4, "Status",               bg=_RED,  size=9)
    for r, s in enumerate(be.get("scenarios", []), start=14):
        bg  = _LGREY if r % 2 == 0 else _WHITE
        col = _GREEN if s["status"] == "safe" else _RED
        _cell(ws, r, 2, f"+{s['cost_increase_pct']}%", bg=bg, align="center")
        _cell(ws, r, 3, s["new_breakeven"], fmt='#,##0', bg=bg, align="right")
        _cell(ws, r, 4, s["status"].upper(), bold=True, color=col, bg=bg)

    # AI insight
    note_row = 20
    ws.merge_cells(f"A{note_row}:D{note_row+1}")
    n = ws[f"A{note_row}"]
    n.value     = f"AI Insight: {be.get('ai_insight', '')}"
    n.font      = Font(name="Arial", italic=True, size=9, color="475569")
    n.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height   = 36
    ws.row_dimensions[note_row+1].height = 36

    _set_col_widths(ws, {"A": 28, "B": 18, "C": 20, "D": 14})


def export_excel_report(
    df, pnl, score, label, alerts, runway_months,
    forecast_data=None, anomaly_data=None, breakeven_data=None,
) -> bytes:
    """
    Build a fully-formatted multi-sheet Excel report.
    Returns raw bytes ready for st.download_button.
    """
    wb = Workbook()
    # Remove default blank sheet
    wb.remove(wb.active)

    _sheet_summary(wb, pnl, score, label, alerts, runway_months)
    _sheet_pnl(wb, df, pnl)
    if forecast_data and "error" not in forecast_data:
        _sheet_forecast(wb, forecast_data)
    if anomaly_data:
        _sheet_anomalies(wb, anomaly_data)
    if breakeven_data:
        _sheet_breakeven(wb, breakeven_data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════
# WEEK 3 — FEATURE 2: PERSISTENT AI CFO CHAT (SUPABASE)
# Supabase table DDL (run once in SQL editor):
#   CREATE TABLE chat_history (
#     id uuid DEFAULT gen_random_uuid() PRIMARY KEY,
#     user_id uuid NOT NULL,
#     role text NOT NULL,
#     content text NOT NULL,
#     session_label text DEFAULT 'default',
#     created_at timestamptz DEFAULT now()
#   );
#   CREATE INDEX ON chat_history (user_id, created_at DESC);
# ═════════════════════════════════════════════════════════

def save_chat_message(db, user_id: str, role: str, content: str,
                      session_label: str = "default") -> bool:
    """Persist a single chat message to Supabase. Returns True on success."""
    try:
        db.table("chat_history").insert({
            "user_id":       user_id,
            "role":          role,
            "content":       content,
            "session_label": session_label,
        }).execute()
        return True
    except Exception:
        return False


def load_chat_history(db, user_id: str, limit: int = 30,
                      session_label: str = "default") -> list:
    """
    Load last `limit` messages for a user from Supabase.
    Returns list of {"role": ..., "content": ...} dicts.
    """
    try:
        result = (
            db.table("chat_history")
            .select("role, content, created_at")
            .eq("user_id", user_id)
            .eq("session_label", session_label)
            .order("created_at", desc=False)
            .limit(limit)
            .execute()
        )
        return [{"role": r["role"], "content": r["content"]} for r in result.data]
    except Exception:
        return []


def build_chat_context_from_history(history: list, max_messages: int = 10) -> list:
    """
    Convert raw history list into Groq messages format,
    prepended with a persistent CFO system prompt.
    """
    system_msg = {
        "role":    "system",
        "content": (
            "You are an expert CFO and business intelligence advisor for an African SME. "
            "You have memory of past conversations — reference them when relevant. "
            "Be concise, cite specific numbers, and give actionable advice."
        ),
    }
    # Take the last max_messages to keep within token limits
    trimmed = history[-max_messages:] if len(history) > max_messages else history
    return [system_msg] + trimmed


def clear_chat_history(db, user_id: str, session_label: str = "default") -> bool:
    """Delete all chat history for a user/session."""
    try:
        db.table("chat_history").delete() \
          .eq("user_id", user_id) \
          .eq("session_label", session_label) \
          .execute()
        return True
    except Exception:
        return False


# ═════════════════════════════════════════════════════════
# WEEK 3 — FEATURE 3: EMAIL REPORT SENDER
# Set env vars: SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS
# Supabase table DDL (run once):
#   CREATE TABLE report_subscriptions (
#     id uuid DEFAULT gen_random_uuid() PRIMARY KEY,
#     user_id uuid NOT NULL UNIQUE,
#     email text NOT NULL,
#     frequency text DEFAULT 'weekly',
#     active boolean DEFAULT true,
#     created_at timestamptz DEFAULT now(),
#     last_sent_at timestamptz
#   );
# ═════════════════════════════════════════════════════════

def send_report_email(
    recipient_email: str,
    pdf_bytes: bytes,
    subject: str = "AI-BOS Weekly Intelligence Report",
    body_html: str = None,
) -> tuple[bool, str]:
    """
    Send the PDF report via SMTP.
    Returns (success: bool, message: str).
    Required env vars: SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS
    """
    smtp_host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.environ.get("SMTP_PORT", 587))
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASS", "")

    if not smtp_user or not smtp_pass:
        return False, "SMTP credentials not configured (set SMTP_USER and SMTP_PASS env vars)."

    if body_html is None:
        body_html = f"""
        <html><body style="font-family:Arial,sans-serif;background:#f8fafc;padding:32px;">
          <div style="max-width:560px;margin:0 auto;background:#fff;border-radius:12px;
                      box-shadow:0 4px 24px rgba(0,0,0,.08);overflow:hidden;">
            <div style="background:#1E3A5F;padding:24px 32px;">
              <h1 style="color:#fff;margin:0;font-size:22px;font-weight:800;">AI-BOS</h1>
              <p style="color:#93C5FD;margin:4px 0 0;font-size:11px;
                        letter-spacing:.1em;text-transform:uppercase;">
                Weekly Intelligence Report</p>
            </div>
            <div style="padding:28px 32px;">
              <p style="color:#334155;font-size:14px;line-height:1.7;">
                Your weekly AI-BOS financial intelligence report is attached.
                Open the Excel workbook for full P&amp;L data, forecasts, anomaly alerts,
                and breakeven analysis.
              </p>
              <p style="color:#64748B;font-size:12px;margin-top:24px;">
                Generated {datetime.date.today().strftime('%A, %d %B %Y')} ·
                AI-BOS Intelligence Platform
              </p>
            </div>
          </div>
        </body></html>
        """

    msg = MIMEMultipart("mixed")
    msg["From"]    = smtp_user
    msg["To"]      = recipient_email
    msg["Subject"] = subject

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(body_html, "html"))
    msg.attach(alt)

    if pdf_bytes:
        att = MIMEBase("application", "pdf")
        att.set_payload(pdf_bytes)
        encoders.encode_base64(att)
        att.add_header("Content-Disposition", "attachment",
                       filename=f"aibos_report_{datetime.date.today()}.pdf")
        msg.attach(att)

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, recipient_email, msg.as_string())
        return True, f"Report sent to {recipient_email}"
    except Exception as e:
        return False, f"Email failed: {e}"


def upsert_subscription(db, user_id: str, email: str,
                        frequency: str = "weekly", active: bool = True) -> bool:
    """Add or update an email subscription in Supabase."""
    try:
        db.table("report_subscriptions").upsert({
            "user_id":   user_id,
            "email":     email,
            "frequency": frequency,
            "active":    active,
        }).execute()
        return True
    except Exception:
        return False


def get_subscription(db, user_id: str) -> dict | None:
    """Return subscription row for a user, or None."""
    try:
        result = (db.table("report_subscriptions")
                    .select("*").eq("user_id", user_id).limit(1).execute())
        return result.data[0] if result.data else None
    except Exception:
        return None


def get_all_active_subscribers(db) -> list:
    """Return all active subscribers (used by the scheduler)."""
    try:
        result = (db.table("report_subscriptions")
                    .select("user_id, email, frequency")
                    .eq("active", True).execute())
        return result.data or []
    except Exception:
        return []


def mark_report_sent(db, user_id: str) -> None:
    """Update last_sent_at timestamp after successful send."""
    try:
        db.table("report_subscriptions").update({
            "last_sent_at": pd.Timestamp.utcnow().isoformat(),
        }).eq("user_id", user_id).execute()
    except Exception:
        pass


# ═════════════════════════════════════════════════════════
# FUNCTION 5 — PLAIN-TEXT AI ANALYSIS (restored)
# ═════════════════════════════════════════════════════════

def get_ai_analysis(pnl: dict, alerts: list) -> str:
    """
    Returns 3 plain-English actionable recommendations as a
    single string. Used for quick chat context and CLI output.
    """
    prompt = f"""
You are an expert CFO advising a small African business.

Business financial data:
- Total Revenue:  K{pnl['total_revenue']:,}
- Total Costs:    K{pnl['total_costs']:,}
- Total Profit:   K{pnl['total_profit']:,}
- Average Margin: {pnl['avg_margin']}%
- Best Month:     {pnl['best_month']}
- Worst Month:    {pnl['worst_month']}
- Alerts:         {len(alerts)} variance detected

Give exactly 3 specific actionable recommendations.
Reference the actual numbers above.
Plain English only. No jargon.
Maximum 2 sentences each.
"""
    try:
        response = client_ai.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            timeout=30,
        )
        return response.choices[0].message.content
    except Exception:
        return (
            f"1. Review {pnl['worst_month']} costs — margin dropped below average.\n"
            f"2. Replicate {pnl['best_month']} conditions to sustain revenue of K{pnl['total_revenue']:,}.\n"
            f"3. Address {len(alerts)} variance alert(s) before the next reporting period."
        )


# ═════════════════════════════════════════════════════════
# RUN BLOCK — test the full engine from the terminal:
#   python engine.py
# Requires test_data.csv with month, revenue, costs columns
# ═════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys

    CSV_PATH = "test_data.csv"
    try:
        _df = pd.read_csv(CSV_PATH)
    except FileNotFoundError:
        print(f"[ERROR] {CSV_PATH} not found — create it with month, revenue, costs columns.")
        sys.exit(1)

    _df["profit"]     = _df["revenue"] - _df["costs"]
    _df["margin_pct"] = (_df["profit"] / _df["revenue"].replace(0, pd.NA) * 100).fillna(0).round(1)

    print("\n══════════════════════════════════════════════════")
    print("  AI-BOS ENGINE — FULL DIAGNOSTIC RUN")
    print("══════════════════════════════════════════════════")

    # ── 1. P&L ────────────────────────────────────────────
    _pnl    = analyse_pnl(_df)
    _alerts = detect_variances(_df)
    _score, _label = health_score(_pnl, _alerts)

    print(f"\n── P&L SUMMARY ──────────────────────────────────")
    print(f"  Health Score : {_score}/100 — {_label}")
    print(f"  Total Revenue: K{_pnl['total_revenue']:,}")
    print(f"  Total Costs  : K{_pnl['total_costs']:,}")
    print(f"  Net Profit   : K{_pnl['total_profit']:,}")
    print(f"  Avg Margin   : {_pnl['avg_margin']}%")
    print(f"  Best Month   : {_pnl['best_month']}")
    print(f"  Worst Month  : {_pnl['worst_month']}")

    # ── 2. Cash Flow Forecast ─────────────────────────────
    _cf = forecast_cashflow(_df)
    print(f"\n── CASH FLOW FORECAST (30/60/90 days) ───────────")
    for item in _cf:
        print(f"  Month +{item['month_ahead']}: K{item['projected_cash']:,}  {item['status']}")

    # ── 3. Variance Alerts ────────────────────────────────
    print(f"\n── VARIANCE ALERTS ({len(_alerts)} detected) ────────────────")
    if not _alerts:
        print("  No significant variances.")
    for a in _alerts:
        arrow = "▼" if a["direction"] == "drop" else "▲"
        print(f"  {arrow} {a['month']} — {a['change_pct']:+.1f}%")

    # ── 4. Revenue Forecast ───────────────────────────────
    _fc = forecast_revenue(_df)
    if "error" in _fc:
        print(f"\n── REVENUE FORECAST — {_fc['error']}")
    else:
        print(f"\n── REVENUE FORECAST (linear regression) ─────────")
        print(f"  Trend: {_fc['trend']}  |  Growth: {_fc['growth_rate']:+.1f}%/mo"
              f"  |  Confidence: {_fc['confidence']}/100  |  R2: {_fc['r_squared']}")
        for pt in _fc["forecast"]:
            print(f"  {pt['month']}: K{pt['predicted']:,}  (K{pt['low']:,} – K{pt['high']:,})")
        print(f"  AI: {_fc['ai_explanation']}")

    # ── 5. Anomaly Detection ──────────────────────────────
    _anomalies = detect_anomalies(_df)
    print(f"\n── ANOMALY DETECTION ({len(_anomalies)} found) ──────────────────")
    if not _anomalies:
        print("  No anomalies at default z-threshold (2.0).")
    for a in _anomalies:
        arrow = "▼" if a["direction"] == "drop" else "▲"
        print(f"  [{a['severity'].upper():8s}] {a['month']} · {a['metric']} "
              f"{arrow} {a['change_pct']:+.1f}%  (z={a['z_score']}sigma)")
        print(f"             {a['root_cause']}")

    # ── 6. Breakeven Analysis ─────────────────────────────
    _be = calculate_breakeven(_df)
    _be_status = "ABOVE" if _be["current_avg_revenue"] >= _be["breakeven_revenue"] else "BELOW"
    print(f"\n── BREAKEVEN ANALYSIS ───────────────────────────")
    print(f"  Status              : {_be_status} breakeven")
    print(f"  Breakeven Revenue   : K{_be['breakeven_revenue']:,}/month")
    print(f"  Current Avg Revenue : K{_be['current_avg_revenue']:,}/month")
    print(f"  Margin of Safety    : {_be['margin_of_safety_pct']}%  (K{_be['margin_of_safety']:,})")
    print(f"  Contribution Margin : {_be['contribution_margin_ratio']}%")
    print("  What-if scenarios:")
    for s in _be["scenarios"]:
        print(f"    +{s['cost_increase_pct']}% costs -> BE: K{s['new_breakeven']:,}  [{s['status'].upper()}]")
    print(f"  AI: {_be['ai_insight']}")

    # ── 7. Plain-text AI Recommendations ─────────────────
    print(f"\n── AI RECOMMENDATIONS (plain text) ──────────────")
    _recs = get_ai_analysis(_pnl, _alerts)
    for line in _recs.strip().split("\n"):
        print(f"  {line}")

    # ── 8. Structured JSON Recommendations ───────────────
    print(f"\n── STRUCTURED RECOMMENDATIONS (JSON) ────────────")
    _structured = get_structured_analysis(_pnl, _alerts)
    for i, r in enumerate(_structured, 1):
        print(f"  [{r.get('priority','?').upper():6s}] REC {i}: {r.get('title','')}")
        print(f"           {r.get('recommendation','')}")

    # ── 9. Excel Export (writes file to disk) ────────────
    print(f"\n── EXCEL EXPORT ─────────────────────────────────")
    try:
        _xlsx = export_excel_report(
            _df, _pnl, _score, _label, _alerts,
            runway_months=0.0,
            forecast_data=_fc if "error" not in _fc else None,
            anomaly_data=_anomalies,
            breakeven_data=_be,
        )
        out_path = "aibos_test_report.xlsx"
        with open(out_path, "wb") as fh:
            fh.write(_xlsx)
        print(f"  Written to {out_path}  ({len(_xlsx):,} bytes)")
    except Exception as exc:
        print(f"  Excel export failed: {exc}")

    
def export_excel_report(df, pnl, health_score, health_label, alerts,
                        runway_months, forecast_data=None,
                        anomaly_data=None, breakeven_data=None) -> bytes:
    import io
    import pandas as pd
    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["AI-BOS Intelligence Report"])
    ws.append([])
    ws.append(["Health Score", f"{health_score}/100 ({health_label})"])
    ws.append(["Total Revenue", pnl.get("total_revenue", 0)])
    ws.append(["Total Costs",   pnl.get("total_costs",   0)])
    ws.append(["Total Profit",  pnl.get("total_profit",  0)])
    ws.append(["Avg Margin %",  pnl.get("avg_margin",    0)])
    ws.append(["Cash Runway",   f"{runway_months:.1f} months"])
    ws.append(["Best Month",    pnl.get("best_month",  "")])
    ws.append(["Worst Month",   pnl.get("worst_month", "")])

    ws2 = wb.create_sheet("P&L Data")
    cols = [c for c in ["month","revenue","costs","profit","margin_pct"] if c in df.columns]
    ws2.append(cols)
    for row in df[cols].itertuples(index=False):
        ws2.append(list(row))

    ws3 = wb.create_sheet("Alerts")
    ws3.append(["Month", "Direction", "Change %", "Type"])
    for a in alerts:
        ws3.append([a.get("month",""), a.get("direction",""),
                    a.get("change_pct", 0), a.get("type","")])

    if forecast_data and "forecast" in forecast_data:
        ws4 = wb.create_sheet("Forecast")
        ws4.append(["Month", "Predicted", "Low", "High"])
        for pt in forecast_data["forecast"]:
            ws4.append([pt.get("month",""), pt.get("predicted",0),
                        pt.get("low",0),  pt.get("high",0)])

    if breakeven_data:
        ws5 = wb.create_sheet("Breakeven")
        ws5.append(["Metric", "Value"])
        for k, v in breakeven_data.items():
            if not isinstance(v, (list, dict)):
                ws5.append([k, v])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

    print("\n══════════════════════════════════════════════════")
    print("  ENGINE RUN COMPLETE")
    print("══════════════════════════════════════════════════\n")